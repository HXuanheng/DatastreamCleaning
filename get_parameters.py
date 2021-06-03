import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook


in_folder ="resources/"
out_folder="results/"

def readsheets(fname):
    xlsx = pd.ExcelFile(in_folder+fname, engine='openpyxl')
    sheet_names = xlsx.sheet_names
    dfs = pd.read_excel(in_folder+fname, sheet_name=None, engine='openpyxl')
    writer = pd.ExcelWriter(out_folder+fname, engine='xlsxwriter')
    for sheet_name in sheet_names:
        dfs[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()

def openpyxl(fname):
    wb = load_workbook(fname)
    return wb

def readcomments(wb,sheet_name,cell):
    ws = wb[sheet_name]
    comment = ws[cell].comment
    comment = str(comment)
    return comment

# get parameter from comment
def getparameters(wb,sheet_name,cell):
    comment = readcomments(wb,sheet_name,cell)
    match = re.findall(r'\"(.*?)\"', comment)
    return match

def vardetail(fname):
    xlsx = pd.ExcelFile(in_folder+fname, engine='openpyxl')
    sheet_names = xlsx.sheet_names
    fields = []
    freqs = []
    timeseries = []
    var_names = []
    wb = openpyxl(in_folder+fname)
    for sheet_name in sheet_names:
        match = getparameters(wb,sheet_name,"A1")
        field = match[0]
        freq = match[3]
        if not freq:
            timeserie = ""
        else:
            timeserie = "Y"
        var_names.append(sheet_name)
        fields.append(field)
        freqs.append(freq)
        timeseries.append(timeserie)
        print(sheet_name + ' - done!')
    df = pd.DataFrame({'Datastream data': fields,
                  'Frequency': freqs,
                  'time series': timeseries,
                  'Variable name': var_names})
    
    df.to_excel(out_folder+fname, index=False)

if __name__ == "__main__":
    for n, filename in enumerate(os.listdir(in_folder)):
        print("({})Processing {}...".format(str(n),filename))
        if filename != '.DS_Store':
            vardetail(filename)
    print('Process Completed...')