import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import time
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows

#fields = "Data - Daily.xlsx"
#fields = "Data - Monthly.xlsx"
#fields = "Data - Yearly.xlsx"
#fields = "Data - LatestValue.xlsx"
fields = "fields.xlsx"


#id_company = "list_identifiers.xlsx"
id_company = "identifiers.xlsx"

in_folder ="resources/parameters/"
out_folder="results/"

def get_parameter(fields,fid):
    df = pd.read_excel(in_folder+fields, engine='openpyxl')
    try:
        df['end date'] = pd.to_datetime(df['end date'], format='%Y-%m-%d %H:%M:%S')
        df['start date'] = pd.to_datetime(df['start date'], format='%Y-%m-%d %H:%M:%S')
        df['end date'] = df['end date'].dt.strftime('%Y-%m-%d')
        df['start date'] = df['start date'].dt.strftime('%Y-%m-%d')
    except:
        pass
    df = df.replace(np.nan, '', regex=True)
    varcode = df['Datastream data']
    freq = df['Frequency']
    sdate = df['start date']
    edate = df['end date']
    varname = df['Variable name']
    df_id = pd.read_excel(in_folder+fid, engine='openpyxl')
    ncompany = len(df_id.index)
    return varcode, freq, sdate, edate, varname, ncompany

def write_excel(fid, varcode, freq, sdate, edate, varname, ncompany):
    df = pd.read_excel(in_folder+fid, engine='openpyxl')
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=None, header=True):
        ws.append(r)
    for n, x in enumerate(varname):
        ws = wb.create_sheet(x)
        if 'Value' in sdate[n]:
            formula = '=DSGRID(\'Sheet\'!$A$2:$A$' + str(ncompany+1) + ',"' + str(varcode[n]) + '","' + str(sdate[n]) + '","' + str(edate[n]) + '","' + str(freq[n]) + '","RowHeader=true;ColHeader=true;Heading=true;Transpose=true;Curn=true;DispSeriesDescription=true;DispDatatypeDescription=true","")'
        else:    
            formula = '=DSGRID(\'Sheet\'!$A$2:$A$' + str(ncompany+1) + ',"' + str(varcode[n]) + '","' + str(sdate[n]) + '","' + str(edate[n]) + '","' + str(freq[n]) + '","RowHeader=true;ColHeader=true;Heading=true;Code=true;Curn=true;DispSeriesDescription=false;YearlyTSFormat=false;QuarterlyTSFormat=false","")'
        ws['A1'] = formula
    if str(freq[n]) == '':
        wb.save(out_folder + 'Data_freq_' + str(sdate[n]) + '.xlsx')
    else:    
        wb.save(out_folder + 'Data_freq_' + str(freq[n]) + '.xlsx')


if __name__ == "__main__":
    varcode, freq, sdate, edate, varname, ncompany = get_parameter(fields, id_company)
    write_excel(id_company, varcode, freq, sdate, edate, varname, ncompany)

# =DSGRID('unique firms'!$B$2:$B$327,"WC01801","2008","2018","Y","RowHeader=true;ColHeader=true;Heading=true;Code=true;Curn=true;DispSeriesDescription=false;YearlyTSFormat=false;QuarterlyTSFormat=false","")
# =DSGRID('unique firms'!$B$2:$B$327,"RI","2008","2018","D","RowHeader=true;ColHeader=true;Heading=true;Curn=true;DispSeriesDescription=false;YearlyTSFormat=false;QuarterlyTSFormat=false","")
# =DSGRID('unique firms'!$B$2:$B$327,"INDG","Latest Value","","","RowHeader=true;ColHeader=true;Heading=true;Transpose=true;Curn=true;DispSeriesDescription=true;DispDatatypeDescription=true","")